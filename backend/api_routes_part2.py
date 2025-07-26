#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Routes API complètes - Partie 2
Gestion des notes, calculs de moyennes, génération PDF
"""

from fastapi import APIRouter, HTTPException, Depends, status, Response
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import uuid
import statistics
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import base64
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Import depuis le fichier principal
from .server import (
    db, require_role, RoleEnum, get_current_user,
    GradeCreate, GradeUpdate, GradeResponse, EnrollmentCreate,
    TranscriptResponse, TranscriptStatusEnum
)

# Routeur pour les routes de notes
grades_router = APIRouter()

# ===========================
# ROUTES GESTION NOTES
# ===========================

@grades_router.post("/api/grades", response_model=GradeResponse)
async def create_grade(
    grade_request: GradeCreate,
    current_user: dict = Depends(require_role([RoleEnum.TEACHER, RoleEnum.ADMIN]))
):
    """
    Créer une nouvelle note (saisie par enseignant ou admin)
    Équivalent au contrôleur de création de notes Spring Boot
    """
    # Vérifier que l'étudiant existe
    student = db.users.find_one({"_id": grade_request.student_id, "role": "STUDENT"})
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Étudiant non trouvé"
        )
    
    # Vérifier que la matière existe
    subject = db.subjects.find_one({"_id": grade_request.subject_id})
    if not subject:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Matière non trouvée"
        )
    
    # Créer la note
    grade_id = str(uuid.uuid4())
    grade_data = {
        "_id": grade_id,
        "student_id": grade_request.student_id,
        "subject_id": grade_request.subject_id,
        "value": grade_request.value,
        "comment": grade_request.comment,
        "recorded_by_teacher_id": grade_request.recorded_by_teacher_id or str(current_user["_id"]),
        "date": datetime.utcnow()
    }
    
    db.grades.insert_one(grade_data)
    
    # Préparer la réponse
    teacher_name = ""
    if grade_data["recorded_by_teacher_id"]:
        teacher = db.users.find_one({"_id": grade_data["recorded_by_teacher_id"]})
        if teacher:
            teacher_name = f"{teacher['firstname']} {teacher['lastname']}"
    
    return GradeResponse(
        id=grade_id,
        value=grade_request.value,
        date=datetime.utcnow(),
        comment=grade_request.comment,
        student_id=grade_request.student_id,
        student_name=f"{student['firstname']} {student['lastname']}",
        subject_id=grade_request.subject_id,
        subject_name=subject["name"],
        subject_coefficient=subject["coefficient"],
        recorded_by=teacher_name
    )

@grades_router.get("/api/grades/student/{student_id}", response_model=List[GradeResponse])
async def get_student_grades(
    student_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Obtenir toutes les notes d'un étudiant
    Accessible à l'étudiant lui-même, ses enseignants ou l'admin
    """
    # Vérification des permissions
    if (current_user["role"] == "STUDENT" and str(current_user["_id"]) != student_id and 
        current_user["role"] not in ["TEACHER", "ADMIN"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès non autorisé à ces notes"
        )
    
    # Récupérer toutes les notes de l'étudiant
    grades = list(db.grades.find({"student_id": student_id}))
    student = db.users.find_one({"_id": student_id})
    
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Étudiant non trouvé"
        )
    
    result = []
    for grade in grades:
        subject = db.subjects.find_one({"_id": grade["subject_id"]})
        teacher_name = ""
        if grade.get("recorded_by_teacher_id"):
            teacher = db.users.find_one({"_id": grade["recorded_by_teacher_id"]})
            if teacher:
                teacher_name = f"{teacher['firstname']} {teacher['lastname']}"
        
        result.append(GradeResponse(
            id=str(grade["_id"]),
            value=grade["value"],
            date=grade["date"],
            comment=grade.get("comment"),
            student_id=student_id,
            student_name=f"{student['firstname']} {student['lastname']}",
            subject_id=grade["subject_id"],
            subject_name=subject["name"] if subject else "Matière inconnue",
            subject_coefficient=subject["coefficient"] if subject else 1.0,
            recorded_by=teacher_name
        ))
    
    return result

@grades_router.get("/api/grades/subject/{subject_id}", response_model=List[GradeResponse])
async def get_subject_grades(
    subject_id: str,
    current_user: dict = Depends(require_role([RoleEnum.TEACHER, RoleEnum.ADMIN]))
):
    """
    Obtenir toutes les notes d'une matière
    Accessible aux enseignants et administrateurs
    """
    grades = list(db.grades.find({"subject_id": subject_id}))
    subject = db.subjects.find_one({"_id": subject_id})
    
    if not subject:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Matière non trouvée"
        )
    
    result = []
    for grade in grades:
        student = db.users.find_one({"_id": grade["student_id"]})
        teacher_name = ""
        if grade.get("recorded_by_teacher_id"):
            teacher = db.users.find_one({"_id": grade["recorded_by_teacher_id"]})
            if teacher:
                teacher_name = f"{teacher['firstname']} {teacher['lastname']}"
        
        result.append(GradeResponse(
            id=str(grade["_id"]),
            value=grade["value"],
            date=grade["date"],
            comment=grade.get("comment"),
            student_id=grade["student_id"],
            student_name=f"{student['firstname']} {student['lastname']}" if student else "Étudiant inconnu",
            subject_id=subject_id,
            subject_name=subject["name"],
            subject_coefficient=subject["coefficient"],
            recorded_by=teacher_name
        ))
    
    return result

@grades_router.put("/api/grades/{grade_id}", response_model=GradeResponse)
async def update_grade(
    grade_id: str,
    grade_update: GradeUpdate,
    current_user: dict = Depends(require_role([RoleEnum.TEACHER, RoleEnum.ADMIN]))
):
    """
    Modifier une note existante
    Accessible aux enseignants et administrateurs
    """
    grade = db.grades.find_one({"_id": grade_id})
    if not grade:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Note non trouvée"
        )
    
    # Préparer les modifications
    update_data = {}
    if grade_update.value is not None:
        update_data["value"] = grade_update.value
    if grade_update.comment is not None:
        update_data["comment"] = grade_update.comment
    
    update_data["modified_at"] = datetime.utcnow()
    update_data["modified_by"] = str(current_user["_id"])
    
    # Effectuer la mise à jour
    db.grades.update_one({"_id": grade_id}, {"$set": update_data})
    
    # Récupérer la note mise à jour pour la réponse
    updated_grade = db.grades.find_one({"_id": grade_id})
    student = db.users.find_one({"_id": updated_grade["student_id"]})
    subject = db.subjects.find_one({"_id": updated_grade["subject_id"]})
    
    teacher_name = ""
    if updated_grade.get("recorded_by_teacher_id"):
        teacher = db.users.find_one({"_id": updated_grade["recorded_by_teacher_id"]})
        if teacher:
            teacher_name = f"{teacher['firstname']} {teacher['lastname']}"
    
    return GradeResponse(
        id=grade_id,
        value=updated_grade["value"],
        date=updated_grade["date"],
        comment=updated_grade.get("comment"),
        student_id=updated_grade["student_id"],
        student_name=f"{student['firstname']} {student['lastname']}" if student else "Étudiant inconnu",
        subject_id=updated_grade["subject_id"],
        subject_name=subject["name"] if subject else "Matière inconnue",
        subject_coefficient=subject["coefficient"] if subject else 1.0,
        recorded_by=teacher_name
    )

@grades_router.delete("/api/grades/{grade_id}")
async def delete_grade(
    grade_id: str,
    current_user: dict = Depends(require_role([RoleEnum.TEACHER, RoleEnum.ADMIN]))
):
    """
    Supprimer une note
    Accessible aux enseignants et administrateurs
    """
    result = db.grades.delete_one({"_id": grade_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Note non trouvée"
        )
    
    return {"message": "Note supprimée avec succès"}

# ===========================
# CALCULS DE MOYENNES
# ===========================

@grades_router.get("/api/students/{student_id}/average")
async def calculate_student_average(
    student_id: str,
    semester: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Calculer la moyenne générale pondérée d'un étudiant
    Équivalent aux calculs de moyenne du backend Spring Boot
    """
    # Vérification des permissions
    if (current_user["role"] == "STUDENT" and str(current_user["_id"]) != student_id and 
        current_user["role"] not in ["TEACHER", "ADMIN"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès non autorisé"
        )
    
    # Récupérer toutes les notes de l'étudiant
    grades = list(db.grades.find({"student_id": student_id}))
    
    if not grades:
        return {
            "student_id": student_id,
            "general_average": 0,
            "subject_averages": [],
            "total_coefficient": 0,
            "semester": semester
        }
    
    # Grouper les notes par matière
    subjects_data = {}
    for grade in grades:
        subject_id = grade["subject_id"]
        if subject_id not in subjects_data:
            subject = db.subjects.find_one({"_id": subject_id})
            if subject:
                subjects_data[subject_id] = {
                    "subject_name": subject["name"],
                    "coefficient": subject["coefficient"],
                    "grades": []
                }
        
        if subject_id in subjects_data:
            subjects_data[subject_id]["grades"].append(grade["value"])
    
    # Calculer les moyennes par matière et la moyenne générale
    subject_averages = []
    weighted_sum = 0
    total_coefficient = 0
    
    for subject_id, data in subjects_data.items():
        if data["grades"]:
            subject_avg = statistics.mean(data["grades"])
            subject_averages.append({
                "subject_id": subject_id,
                "subject_name": data["subject_name"],
                "average": round(subject_avg, 2),
                "coefficient": data["coefficient"],
                "grade_count": len(data["grades"])
            })
            
            # Calcul pondéré pour la moyenne générale
            weighted_sum += subject_avg * data["coefficient"]
            total_coefficient += data["coefficient"]
    
    general_average = round(weighted_sum / total_coefficient, 2) if total_coefficient > 0 else 0
    
    return {
        "student_id": student_id,
        "general_average": general_average,
        "subject_averages": subject_averages,
        "total_coefficient": total_coefficient,
        "semester": semester,
        "calculation_date": datetime.utcnow()
    }

# ===========================
# GÉNÉRATION DE RELEVÉS PDF
# ===========================

@grades_router.get("/api/students/{student_id}/transcript/pdf")
async def generate_transcript_pdf(
    student_id: str,
    semester: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Générer un relevé de notes PDF
    Équivalent à la génération PDF du backend Spring Boot
    """
    # Vérification des permissions
    if (current_user["role"] == "STUDENT" and str(current_user["_id"]) != student_id and 
        current_user["role"] not in ["TEACHER", "ADMIN"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès non autorisé"
        )
    
    # Récupérer les données de l'étudiant
    student = db.users.find_one({"_id": student_id})
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Étudiant non trouvé"
        )
    
    # Récupérer les notes et calculer les moyennes
    average_data = await calculate_student_average(student_id, semester, current_user)
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, 
                           topMargin=72, bottomMargin=18)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Contenu du PDF
    story = []
    
    # Titre
    story.append(Paragraph("RELEVÉ DE NOTES", title_style))
    story.append(Spacer(1, 12))
    
    # Informations étudiant
    student_info = [
        ["Nom :", f"{student['lastname']} {student['firstname']}"],
        ["N° Étudiant :", student.get('student_id_num', 'N/A')],
        ["Email :", student['email']],
        ["Date d'édition :", datetime.now().strftime("%d/%m/%Y")]
    ]
    
    if semester:
        student_info.append(["Semestre :", semester])
    
    student_table = Table(student_info, colWidths=[2*inch, 3*inch])
    student_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(student_table)
    story.append(Spacer(1, 20))
    
    # Tableau des notes par matière
    notes_data = [["Matière", "Coefficient", "Nombre de notes", "Moyenne"]]
    
    for subject in average_data["subject_averages"]:
        notes_data.append([
            subject["subject_name"],
            str(subject["coefficient"]),
            str(subject["grade_count"]),
            f"{subject['average']}/20"
        ])
    
    # Ligne de moyenne générale
    notes_data.append(["", "", "", ""])  # Ligne vide
    notes_data.append(["MOYENNE GÉNÉRALE", "", "", f"{average_data['general_average']}/20"])
    
    notes_table = Table(notes_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    notes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -3), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(notes_table)
    
    # Construire le PDF
    doc.build(story)
    
    # Préparer la réponse
    buffer.seek(0)
    pdf_content = buffer.read()
    buffer.close()
    
    # Encoder en base64 pour la réponse JSON
    pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
    
    # Enregistrer le transcript en base
    transcript_id = str(uuid.uuid4())
    transcript_data = {
        "_id": transcript_id,
        "student_id": student_id,
        "generation_date": datetime.utcnow(),
        "status": TranscriptStatusEnum.GENERATED.value,
        "filepath": f"transcripts/{transcript_id}.pdf",
        "semester": semester
    }
    db.transcripts.insert_one(transcript_data)
    
    return {
        "transcript_id": transcript_id,
        "student_id": student_id,
        "pdf_base64": pdf_base64,
        "filename": f"releve_notes_{student['lastname']}_{student['firstname']}.pdf",
        "generation_date": datetime.utcnow()
    }

# ===========================
# EXPORT EXCEL
# ===========================

@grades_router.get("/api/admin/export/excel")
async def export_grades_excel(
    current_user: dict = Depends(require_role([RoleEnum.ADMIN]))
):
    """
    Exporter toutes les notes en Excel
    Accessible uniquement aux administrateurs
    """
    # Créer un nouveau workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes"
    
    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = ["Étudiant", "N° Étudiant", "Matière", "Note", "Coefficient", 
               "Date", "Commentaire", "Enseignant"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer toutes les notes avec les informations associées
    grades = list(db.grades.find({}))
    
    for row, grade in enumerate(grades, 2):
        student = db.users.find_one({"_id": grade["student_id"]})
        subject = db.subjects.find_one({"_id": grade["subject_id"]})
        teacher = None
        if grade.get("recorded_by_teacher_id"):
            teacher = db.users.find_one({"_id": grade["recorded_by_teacher_id"]})
        
        ws.cell(row=row, column=1, value=f"{student['lastname']} {student['firstname']}" if student else "N/A")
        ws.cell(row=row, column=2, value=student.get('student_id_num', 'N/A') if student else "N/A")
        ws.cell(row=row, column=3, value=subject['name'] if subject else "N/A")
        ws.cell(row=row, column=4, value=grade['value'])
        ws.cell(row=row, column=5, value=subject['coefficient'] if subject else 1.0)
        ws.cell(row=row, column=6, value=grade['date'].strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=7, value=grade.get('comment', ''))
        ws.cell(row=row, column=8, value=f"{teacher['firstname']} {teacher['lastname']}" if teacher else "")
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_content = buffer.read()
    buffer.close()
    
    # Encoder en base64
    excel_base64 = base64.b64encode(excel_content).decode('utf-8')
    
    return {
        "excel_base64": excel_base64,
        "filename": f"export_notes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "export_date": datetime.utcnow(),
        "total_grades": len(grades)
    }